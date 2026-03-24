import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

EXCEL_FILE = "Attendance_Data.xlsx"

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        # Keep one sheet minimum or add before removing
        wb.save(EXCEL_FILE)

def get_sheet(class_name):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    if class_name not in wb.sheetnames:
        ws = wb.create_sheet(class_name)
        ws.cell(row=1, column=1, value="Roll No")
        wb.save(EXCEL_FILE)
    else:
        ws = wb[class_name]
    return wb, ws

def add_attendance(class_name, date_str, roll_no):
    wb, ws = get_sheet(class_name)
    
    # 1. Find or Create Date Column
    date_col = None
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == date_str:
            date_col = col
            break
    
    if not date_col:
        # Check if last column is "Percentage (%)"
        last_col = ws.max_column
        if ws.cell(row=1, column=last_col).value == "Percentage (%)":
             # We need to insert the date before the percentage column
             ws.insert_cols(last_col)
             date_col = last_col
        else:
            date_col = ws.max_column + 1
        ws.cell(row=1, column=date_col, value=date_str)

    # 2. Find or Create Roll No Row
    roll_row = None
    all_rolls = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            all_rolls.append((val, row))
            if str(val) == str(roll_no):
                roll_row = row

    if not roll_row:
        # Add new roll no and sort
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=int(roll_no))
        ws.cell(row=new_row, column=date_col, value="P")
        wb.save(EXCEL_FILE)
        sort_roll_numbers(class_name)
    else:
        # Mark P
        ws.cell(row=roll_row, column=date_col, value="P")
        wb.save(EXCEL_FILE)

def delete_attendance(class_name, date_str, roll_no):
    wb, ws = get_sheet(class_name)
    
    date_col = None
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == date_str:
            date_col = col
            break
            
    if not date_col:
        return False

    roll_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(roll_no):
            roll_row = row
            break
    
    if roll_row:
        ws.cell(row=roll_row, column=date_col, value="")
        wb.save(EXCEL_FILE)
        return True
    return False

def sort_roll_numbers(class_name):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[class_name]
    
    data = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data.append(row)
    
    # Sort by Roll No (column 0)
    data.sort(key=lambda x: int(x[0]))
    
    # Clear and rewrite
    ws.delete_rows(2, ws.max_row)
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    wb.save(EXCEL_FILE)

def calculate_percentage(class_name):
    wb, ws = get_sheet(class_name)
    
    # Find active date columns (exclude Roll No and existing Percentage)
    date_cols = []
    perc_col = None
    
    for col in range(2, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == "Percentage (%)":
            perc_col = col
        elif header:
            date_cols.append(col)
            
    if not date_cols:
        return "No attendance data to calculate."

    if not perc_col:
        perc_col = ws.max_column + 1
        ws.cell(row=1, column=perc_col, value="Percentage (%)")

    total_days = len(date_cols)
    for row in range(2, ws.max_row + 1):
        p_count = 0
        for col in date_cols:
            if ws.cell(row=row, column=col).value == "P":
                p_count += 1
        
        percentage = (p_count / total_days * 100) if total_days > 0 else 0
        ws.cell(row=row, column=perc_col, value=f"{percentage:.2f}%")
    
    wb.save(EXCEL_FILE)
    return "Percentage calculated successfully."

def delete_percentage_column(class_name):
    wb, ws = get_sheet(class_name)
    perc_col = None
    for col in range(2, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Percentage (%)":
            perc_col = col
            break
    
    if perc_col:
        ws.delete_cols(perc_col)
        wb.save(EXCEL_FILE)
        return "Percentage column deleted."
    return "No percentage column found."

def reset_data():
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    ensure_excel_file()

def get_all_classes():
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    return wb.sheetnames

def get_student_stats(class_name, roll_no):
    wb, ws = get_sheet(class_name)
    
    roll_row = None
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == str(roll_no):
            roll_row = row
            break
            
    if not roll_row:
        return None

    date_cols = []
    for col in range(2, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and header != "Percentage (%)":
            date_cols.append(col)
            
    p_count = 0
    for col in date_cols:
        if ws.cell(row=roll_row, column=col).value == "P":
            p_count += 1
            
    total_days = len(date_cols)
    percentage = (p_count / total_days * 100) if total_days > 0 else 0
    
    return {
        "days": p_count,
        "total": total_days,
        "percentage": f"{percentage:.2f}%"
    }

def add_class(class_name):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    if class_name not in wb.sheetnames:
        ws = wb.create_sheet(class_name)
        ws.cell(row=1, column=1, value="Roll No")
        wb.save(EXCEL_FILE)
        return True
    return False

def remove_class(class_name):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    if class_name in wb.sheetnames:
        del wb[class_name]
        wb.save(EXCEL_FILE)
        return True
    return False
